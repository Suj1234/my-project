import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Native Fields Review"

# ── Styles ──────────────────────────────────────────────────────────────────
header_fill   = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
section_fill  = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
error_fill    = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
pending_fill  = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
changed_fill  = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
nochange_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

header_font  = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
section_font = Font(bold=True, color="1F3864", size=11, name="Calibri")
body_font    = Font(size=10, name="Calibri")
error_font   = Font(size=10, name="Calibri", color="C00000")
pending_font = Font(size=10, name="Calibri", color="7F6000")

wrap = Alignment(wrap_text=True, vertical="top")
center_wrap = Alignment(wrap_text=True, vertical="top", horizontal="center")

thin_border = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

# ── Column headers ───────────────────────────────────────────────────────────
HEADERS = [
    "Table",
    "Native Field Key (BE)",
    "Current Display Name",
    "Current Description",
    "Comments",
    "New Display Name",
    "New Description",
]

# ── Data ─────────────────────────────────────────────────────────────────────
# Row type: "section" | "none" | "error" | "pending" | "change"
# Format : (type, table, native_key, curr_display, curr_desc, comments, new_display, new_desc)

ROWS = [
    # ── APPLICANT ─────────────────────────────────────────────────────────────
    ("section", "APPLICANT", "", "", "", "", "", ""),
    ("none",    "applicant", "$native.applicant.id",          "Applicant ID",           "Unique identifier of the applicant",                   "No change needed",                                                                                          "Applicant ID",             "Unique identifier of the applicant"),
    ("change",  "applicant", "$native.applicant.uuid",         "Applicant UUID",         "A unique identifier of the applicant",                  "'UUID' is developer jargon — not meaningful to CST users. Description also nearly identical to id — adds no clarity.", "Applicant Reference Code", "A system-generated unique reference code for the applicant"),
    ("change",  "applicant", "$native.applicant.tenantId",     "Tenant ID",              "Tenant ID associated with the applicant",               "Description restates the display name — adds no new information. Display name updated to match business terminology.",  "Lender ID",                "Unique ID of the lender this applicant is associated with"),
    ("change",  "applicant", "$native.applicant.phoneNo",      "Phone Number",           "Primary contact phone number of the applicant",         "Display name lacks 'Applicant' prefix — in a flat list 'Phone Number' has no context.",                     "Applicant Phone Number",   "Primary contact phone number of the applicant"),
    ("none",    "applicant", "$native.applicant.status",       "Applicant Status",       "Current status of the applicant",                      "No change needed",                                                                                          "Applicant Status",         "Current status of the applicant"),
    ("change",  "applicant", "$native.applicant.dateCreated",  "Date Created",           "Timestamp when the applicant was created",              "'Applicant was created' sounds like the person was created, not the record.",                               "Date Created",             "Timestamp when the applicant record was created"),
    ("change",  "applicant", "$native.applicant.lastUpdated",  "Last Updated",           "Timestamp when the applicant was last updated",         "'Applicant was last updated' is awkward — should reference the record.",                                    "Last Updated",             "Timestamp when the applicant record was last updated"),

    # ── DOCUMENT ──────────────────────────────────────────────────────────────
    ("section", "DOCUMENT", "", "", "", "", "", ""),
    ("none",    "document", "$native.document.id",                    "Document ID",             "Unique identifier of the document",                              "No change needed",                                                                                                            "Document ID",              "Unique identifier of the document"),
    ("change",  "document", "$native.document.tenantId",              "Tenant ID",               "Tenant ID associated with the document",                         "Description restates the display name — adds no new information. Display name updated to match business terminology.",        "Lender ID",                "Unique ID of the lender this document is associated with"),
    ("change",  "document", "$native.document.uuid",                  "Document UUID",           "A unique identifier of the document",                            "'UUID' is developer jargon. Description same as id — no distinction explained.",                                             "Document Reference Code",  "A system-generated unique reference code for the document"),
    ("change",  "document", "$native.document.onboardingApplicationId","Onboarding Application ID","ID of the respective onboarding application",                  "'Respective' is awkward — not natural language. Other tables use 'Reference ID'.",                                          "Onboarding Application ID","Reference ID of the onboarding application this document belongs to"),
    ("pending", "document", "$native.document.documentIdentifier",    "Document Identifier",     "A unique document identifier",                                  "PENDING CLARIFICATION: How does this differ from Document ID and Document Reference Code? Is this an external or partner-system identifier?", "PENDING CLARIFICATION",    "PENDING CLARIFICATION"),
    ("none",    "document", "$native.document.documentType",          "Document Type",           "Type of the document uploaded",                                 "No change needed",                                                                                                            "Document Type",            "Type of the document uploaded"),
    ("none",    "document", "$native.document.documentUrl",           "Document URL",            "URL of the uploaded document",                                  "No change needed",                                                                                                            "Document URL",             "URL of the uploaded document"),
    ("change",  "document", "$native.document.documentCategory",      "Document Category",       "Category to which the document belongs",                        "'Belongs to' is informal — replacing with clearer phrasing.",                                                                 "Document Category",        "The category this document is classified under"),
    ("none",    "document", "$native.document.status",                "Document Status",         "Current status of the document",                                "No change needed",                                                                                                            "Document Status",          "Current status of the document"),
    ("change",  "document", "$native.document.metadata",              "Document Metadata",       "Metadata associated with the document",                         "Description is circular — says 'metadata' to define 'metadata'. Not helpful to a CST user.",                                "Document Metadata",        "Additional technical attributes and details associated with the document"),
    ("error",   "document", "$native.document.dateCreated",           "Last Created",            "Timestamp when the document was first created",                 "ERROR: Display name 'Last Created' is wrong — should be 'Date Created'. 'First created' in description is also redundant.",  "Date Created",             "Timestamp when the document was created"),
    ("none",    "document", "$native.document.lastUpdated",           "Last Updated",            "Timestamp when the document was last updated",                  "No change needed",                                                                                                            "Last Updated",             "Timestamp when the document was last updated"),

    # ── EMPLOYER ──────────────────────────────────────────────────────────────
    ("section", "EMPLOYER", "", "", "", "", "", ""),
    ("none",    "employer", "$native.employer.id",                    "Employer Record ID",      "Unique identifier of the employer details record",               "No change needed",                                                                                                            "Employer Record ID",       "Unique identifier of the employer details record"),
    ("change",  "employer", "$native.employer.tenantId",              "Tenant ID",               "Tenant ID associated with the employer details",                 "Description restates the display name — adds no new information. Display name updated to match business terminology.",        "Lender ID",                "Unique ID of the lender this employer record is associated with"),
    ("change",  "employer", "$native.employer.onboardingApplicationId","Onboarding Application ID","Reference ID of the onboarding application",                  "Clear but can add more context.",                                                                                             "Onboarding Application ID","Reference ID of the onboarding application this employer record is linked to"),
    ("none",    "employer", "$native.employer.employerName",          "Employer Name",           "Name of the employer organization",                             "No change needed",                                                                                                            "Employer Name",            "Name of the employer organization"),
    ("change",  "employer", "$native.employer.employeeId",            "Employee ID",             "Employee identifier within the organization",                   "'Employee identifier' and 'Employee ID' mean the same thing — description adds no new information.",                         "Employee ID",              "The applicant's employee ID within the employer organization"),
    ("none",    "employer", "$native.employer.employerCode",          "Employer Code",           "Internal code representing the employer",                       "No change needed",                                                                                                            "Employer Code",            "Internal code representing the employer"),
    ("change",  "employer", "$native.employer.employerEmail",         "Employer Email",          "Official email address of the employer or HR",                  "Minor: 'HR' is an abbreviation — expanding slightly for clarity.",                                                            "Employer Email",           "Official email address of the employer or HR department"),
    ("change",  "employer", "$native.employer.hrVerificationStatus",  "HR Verification Status",  "Verification status provided by the employer's HR",              "Minor: expanding 'HR' for clarity.",                                                                                          "HR Verification Status",   "Verification status provided by the employer's HR department"),
    ("none",    "employer", "$native.employer.dateCreated",           "Date Created",            "Timestamp when the employer details record was created",         "No change needed",                                                                                                            "Date Created",             "Timestamp when the employer details record was created"),
    ("none",    "employer", "$native.employer.lastUpdated",           "Last Updated",            "Timestamp when the employer details record was last updated",    "No change needed",                                                                                                            "Last Updated",             "Timestamp when the employer details record was last updated"),

    # ── KYC ───────────────────────────────────────────────────────────────────
    ("section", "KYC", "", "", "", "", "", ""),
    ("error",   "kyc", "$native.kyc.id",                             "KYC ID",                  "ID of the KYC table",                                           "ERROR: 'ID of the KYC table' describes the database table, not the record — fundamentally wrong.",                           "KYC Record ID",            "Unique identifier of the KYC record"),
    ("change",  "kyc", "$native.kyc.tenantId",                       "Tenant ID",               "Tenant ID for KYC",                                             "'For KYC' is too brief and inconsistent with all other tables. Display name updated to match business terminology.",           "Lender ID",                "Unique ID of the lender this KYC record is associated with"),
    ("change",  "kyc", "$native.kyc.onboardingApplicationId",        "Onboarding Application ID","ID of the respective onboarding application",                  "'Respective' is awkward — not natural language. Other tables use 'Reference ID'.",                                          "Onboarding Application ID","Reference ID of the onboarding application this KYC record belongs to"),
    ("error",   "kyc", "$native.kyc.poi",                            "Proof Of Identity",       "Use to capture proof of identity",                              "ERROR: 'Use to capture' is an instruction, not a description. 'Of' should be lowercase in title case.",                    "Proof of Identity",        "Document submitted by the applicant as proof of identity"),
    ("error",   "kyc", "$native.kyc.poiType",                        "Proof Of Identity Type",  "Use to mention the type of identity",                           "ERROR: 'Use to mention' is an instruction. 'Type of identity' is vague. 'Of' should be lowercase.",                       "Proof of Identity Type",   "Type of document submitted as proof of identity (e.g. Aadhaar, Passport)"),
    ("change",  "kyc", "$native.kyc.poiMetadata",                    "POI Metadata",            "Metadata associated with Proof of Identity",                    "'POI' abbreviation is inconsistent with parent field which spells it out fully. Description is circular.",                 "Proof of Identity Metadata","Additional technical attributes associated with the proof of identity document"),
    ("error",   "kyc", "$native.kyc.poa",                            "Proof Of Address Document","Use to capture proof of address document",                     "ERROR: 'Use to capture' is an instruction. 'Of' should be lowercase in title case.",                                       "Proof of Address Document","Document submitted by the applicant as proof of address"),
    ("change",  "kyc", "$native.kyc.poaType",                        "POA Type",                "Type of Proof of Address document",                             "'POA' abbreviation is inconsistent with parent field.",                                                                       "Proof of Address Type",    "Type of document submitted as proof of address (e.g. utility bill, bank statement)"),
    ("change",  "kyc", "$native.kyc.poaMetadata",                    "POA Metadata",            "Metadata associated with Proof of Address",                     "'POA' abbreviation is inconsistent with parent field. Description is circular.",                                            "Proof of Address Metadata","Additional technical attributes associated with the proof of address document"),
    ("error",   "kyc", "$native.kyc.pob",                            "Proof of Business Document","Use to capture Proof Of Business Document",                   "ERROR: 'Use to capture' is an instruction. Description nearly repeats the display name.",                                  "Proof of Business Document","Document submitted as proof of business existence or registration"),
    ("change",  "kyc", "$native.kyc.pobType",                        "POB Type",                "Type of Proof of Business document",                            "'POB' abbreviation is inconsistent with parent field.",                                                                       "Proof of Business Type",   "Type of document submitted as proof of business (e.g. GST certificate, trade license)"),
    ("change",  "kyc", "$native.kyc.pobMetadata",                    "POB Metadata",            "Metadata associated with Proof of Business",                    "'POB' abbreviation is inconsistent with parent field. Description is circular.",                                            "Proof of Business Metadata","Additional technical attributes associated with the proof of business document"),
    ("change",  "kyc", "$native.kyc.sourceType",                     "KYC Source Type",         "Source type of the KYC information",                            "Description restates the display name with no added clarity.",                                                                "KYC Source Type",          "The channel or method through which KYC information was collected (e.g. manual entry, digital verification)"),
    ("change",  "kyc", "$native.kyc.lenderRegisteredPhoneNo",        "Lender Registered Phone Number","Phone number registered with lender",                     "Missing article 'the' — grammatically incomplete.",                                                                           "Lender Registered Phone Number","Phone number registered with the lender"),
    ("change",  "kyc", "$native.kyc.businessRegisteredPhoneNo",      "Business Registered Phone Number","Phone number registered for business",                  "Missing 'the'. Also inconsistent with lender field ('registered with' vs 'registered for').",                              "Business Registered Phone Number","Phone number registered for the business"),
    ("change",  "kyc", "$native.kyc.registeredPhoneNo",              "Registered Phone Number", "Primary phone number linked to KYC",                            "Minor: adding 'the' for consistency.",                                                                                        "Registered Phone Number",  "Primary phone number linked to the KYC record"),
    ("pending", "kyc", "$native.kyc.applicantRelationshipType",      "Applicant Relationship Type","Relationship type of the applicant",                         "PENDING CLARIFICATION: Relationship to what or whom? Is this primary vs co-applicant? Or a family relationship type?",      "PENDING CLARIFICATION",    "PENDING CLARIFICATION"),
    ("none",    "kyc", "$native.kyc.applicantFirstName",             "Applicant First Name",    "First name of the applicant",                                   "No change needed",                                                                                                            "Applicant First Name",     "First name of the applicant"),
    ("none",    "kyc", "$native.kyc.applicantMiddleName",            "Applicant Middle Name",   "Middle name of the applicant",                                  "No change needed",                                                                                                            "Applicant Middle Name",    "Middle name of the applicant"),
    ("none",    "kyc", "$native.kyc.applicantLastName",              "Applicant Last Name",     "Last name of the applicant",                                    "No change needed",                                                                                                            "Applicant Last Name",      "Last name of the applicant"),
    ("none",    "kyc", "$native.kyc.applicantFullName",              "Applicant Full Name",     "Full name of the applicant",                                    "No change needed",                                                                                                            "Applicant Full Name",      "Full name of the applicant"),
    ("none",    "kyc", "$native.kyc.applicantGender",                "Applicant Gender",        "Gender of the applicant",                                       "No change needed",                                                                                                            "Applicant Gender",         "Gender of the applicant"),
    ("none",    "kyc", "$native.kyc.applicantDob",                   "Applicant Date of Birth", "Date of birth of the applicant",                                "No change needed",                                                                                                            "Applicant Date of Birth",  "Date of birth of the applicant"),
    ("change",  "kyc", "$native.kyc.applicantImage",                 "Applicant Image",         "Image of the applicant",                                        "'Image' is vague — could mean any image. 'Photograph' is more precise and intuitive.",                                      "Applicant Photograph",     "Photograph of the applicant"),
    ("none",    "kyc", "$native.kyc.applicantEmail",                 "Applicant Email",         "Email address of the applicant",                                "No change needed",                                                                                                            "Applicant Email",          "Email address of the applicant"),
    ("pending", "kyc", "$native.kyc.applicantAddress",               "Applicant Address",       "Address of the applicant",                                      "PENDING CLARIFICATION: Is this the permanent address? Since there is a separate communicationAddress, this field type needs to be specified.", "PENDING CLARIFICATION", "PENDING CLARIFICATION"),
    ("change",  "kyc", "$native.kyc.applicantCity",                  "Applicant City",          "City the applicant belongs to",                                 "'Belongs to' is informal and awkward — not natural phrasing.",                                                                "Applicant City",           "City as per the applicant's address"),
    ("error",   "kyc", "$native.kyc.applicantState",                 "Applicant state",         "State the applicant belongs to",                                "ERROR: 'state' is lowercase — should be title case. 'Belongs to' is informal.",                                             "Applicant State",          "State as per the applicant's address"),
    ("error",   "kyc", "$native.kyc.applicantCountry",               "Applicant country",       "Country the applicant belongs to",                              "ERROR: 'country' is lowercase — should be title case. 'Belongs to' is informal.",                                           "Applicant Country",        "Country as per the applicant's address"),
    ("change",  "kyc", "$native.kyc.applicantPinCode",               "Applicant Pincode",       "Pincode of the applicant",                                      "'Pincode of the applicant' is vague — a pincode belongs to an address, not a person.",                                     "Applicant Pincode",        "PIN code of the applicant's address"),
    ("change",  "kyc", "$native.kyc.applicantCommunicationAddress",  "Applicant Communication Address","Communication Address of the applicant",                "'Address' is incorrectly capitalized mid-sentence in the description.",                                                     "Applicant Communication Address","Communication address of the applicant"),
    ("change",  "kyc", "$native.kyc.applicantCommunicationAddressSource","Source Of Applicant Communication Address","Source of communication Address of the applicant","'Of' is incorrectly capitalized in display name. 'Address' incorrectly capitalized in description.", "Source of Applicant Communication Address","Source from which the applicant's communication address was obtained"),
    ("none",    "kyc", "$native.kyc.companyName",                    "Company Name",            "Name of the company",                                           "No change needed",                                                                                                            "Company Name",             "Name of the company"),
    ("change",  "kyc", "$native.kyc.companyDoi",                     "Date of Incorporation",   "Date of Incorporation of the company",                          "Minor: 'of Incorporation' — 'of' should be lowercase in description for consistency.",                                     "Date of Incorporation",    "Date of incorporation of the company"),
    ("none",    "kyc", "$native.kyc.companyEmail",                   "Company Email",           "Email address of the company",                                  "No change needed",                                                                                                            "Company Email",            "Email address of the company"),
    ("change",  "kyc", "$native.kyc.companyRegisteredAddress",       "Company Registered Address","Address registered by the company",                           "'Registered by the company' sounds like the company filed a document — not that it is the company's registered address.",  "Company Registered Address","Officially registered address of the company"),
    ("error",   "kyc", "$native.kyc.companyBusinessAddress",         "Company Business Address", "Address registered by the Business",                           "ERROR: 'Registered by the Business' is wrong phrasing. 'Business' is incorrectly capitalized.",                           "Company Business Address", "Operational address where the company conducts its business"),
    ("error",   "kyc", "$native.kyc.companyBusinessAddressSource",   "Source Of Company Business Address","Source Of Company Business Address",                  "ERROR: Description is identical to display name — adds zero information. 'Of' also incorrectly capitalized in display name.", "Source of Company Business Address","Source from which the company's business address was obtained"),
    ("change",  "kyc", "$native.kyc.annualIncome",                   "Annual Income",           "Declared annual income",                                        "Clear in KYC context but adding applicant reference makes it more explicit.",                                                "Annual Income",            "Applicant's declared annual income"),
    ("none",    "kyc", "$native.kyc.motherName",                     "Mother's Name",           "Name of applicant's mother",                                    "No change needed",                                                                                                            "Mother's Name",            "Name of the applicant's mother"),
    ("none",    "kyc", "$native.kyc.fatherName",                     "Father's Name",           "Name of applicant's father",                                    "No change needed",                                                                                                            "Father's Name",            "Name of the applicant's father"),
    ("change",  "kyc", "$native.kyc.employmentType",                 "Employment Type",         "Type of employment",                                            "Description too brief — just restates the display name with no added clarity.",                                             "Employment Type",          "The applicant's type of employment (e.g. salaried, self-employed, business owner)"),
    ("change",  "kyc", "$native.kyc.dateCreated",                    "Date Created",            "Record creation timestamp",                                     "Description style is inconsistent with all other tables which use 'Timestamp when the ... was created'.",                   "Date Created",             "Timestamp when the KYC record was created"),
    ("change",  "kyc", "$native.kyc.lastUpdated",                    "Last Updated",            "Record last updated timestamp",                                 "Description style is inconsistent with all other tables which use 'Timestamp when the ... was last updated'.",              "Last Updated",             "Timestamp when the KYC record was last updated"),

    # ── APPLICATION DATA ──────────────────────────────────────────────────────
    ("section", "APPLICATION DATA", "", "", "", "", "", ""),
    ("none",    "applicationData", "$native.applicationData.id",                    "Application Data ID",      "Unique identifier of the application data record",          "No change needed",                                                                                                  "Application Data ID",      "Unique identifier of the application data record"),
    ("change",  "applicationData", "$native.applicationData.tenantId",              "Tenant ID",                "Tenant identifier associated with the application data",     "'Tenant identifier' inconsistent with other tables. Display name updated to match business terminology.",        "Lender ID",                "Unique ID of the lender this application data is associated with"),
    ("change",  "applicationData", "$native.applicationData.onboardingApplicationId","Onboarding Application ID","Reference ID of the onboarding application",               "Minor: adding context.",                                                                                            "Onboarding Application ID","Reference ID of the onboarding application this data belongs to"),
    ("change",  "applicationData", "$native.applicationData.key",                   "Data Key",                 "Key used to identify the application data entry",           "'Data Key' and 'Key used to identify' are technical terms — not intuitive for CST users.",                       "Data Field Name",          "The name or label used to identify this piece of application data"),
    ("change",  "applicationData", "$native.applicationData.value",                 "Data Value",               "Value associated with the application data key",            "Description links back to 'key' which is jargon — not helpful.",                                                "Data Field Value",         "The information stored for this application data field"),
    ("none",    "applicationData", "$native.applicationData.dateCreated",            "Date Created",             "Timestamp when the application data record was created",    "No change needed",                                                                                                  "Date Created",             "Timestamp when the application data record was created"),
    ("none",    "applicationData", "$native.applicationData.lastUpdated",            "Last Updated",             "Timestamp when the application data record was last updated","No change needed",                                                                                                 "Last Updated",             "Timestamp when the application data record was last updated"),

    # ── APPLICATION ───────────────────────────────────────────────────────────
    ("section", "APPLICATION", "", "", "", "", "", ""),
    ("none",    "application", "$native.application.id",                    "Application ID",          "Unique identifier of the onboarding application",           "No change needed",                                                                                                        "Application ID",           "Unique identifier of the onboarding application"),
    ("change",  "application", "$native.application.uuid",                   "Application UUID",        "A unique identifier of the onboarding application",          "'UUID' is developer jargon. Description same as id — no distinction explained.",                                       "Application Reference Code","A system-generated unique reference code for the onboarding application"),
    ("change",  "application", "$native.application.tenantId",               "Tenant ID",               "Tenant ID associated with the application",                  "Description restates the display name — adds no new information. Display name updated to match business terminology.",  "Lender ID",                "Unique ID of the lender this application is associated with"),
    ("change",  "application", "$native.application.purposeId",              "Purpose ID",              "Identifier representing the loan purpose",                   "'Purpose ID' alone is vague — doesn't indicate loan context to a CST user.",                                           "Loan Purpose ID",          "The ID that identifies the loan purpose for this application"),
    ("change",  "application", "$native.application.applicantId",            "Applicant ID",            "Reference ID of the applicant",                              "Minor: adding context.",                                                                                                  "Applicant ID",             "Reference ID of the applicant associated with this application"),
    ("change",  "application", "$native.application.applicantType",          "Applicant Type",          "Type of the applicant",                                      "Description too brief — restates the display name with no added clarity.",                                             "Applicant Type",           "The classification of the applicant (e.g. primary applicant, co-applicant)"),
    ("change",  "application", "$native.application.cifId",                  "CIF ID",                  "Customer Information File identifier",                       "CIF ID confirmed as understood by CST users. Minor: 'identifier' changed to 'ID' for consistency.",                  "CIF ID",                   "Customer Information File (CIF) ID associated with this application"),
    ("none",    "application", "$native.application.appliedLoanAmount",      "Applied Loan Amount",     "Loan amount requested by the applicant",                     "No change needed",                                                                                                        "Applied Loan Amount",      "Loan amount requested by the applicant"),
    ("none",    "application", "$native.application.approvedLoanAmount",     "Approved Loan Amount",    "Loan amount approved by the lender",                         "No change needed",                                                                                                        "Approved Loan Amount",     "Loan amount approved by the lender"),
    ("change",  "application", "$native.application.sanctionedLoanAmount",   "Sanctioned Loan Amount",  "Final sanctioned loan amount",                               "Minor: adding subject for clarity.",                                                                                      "Sanctioned Loan Amount",   "Final loan amount sanctioned for the applicant"),
    ("none",    "application", "$native.application.status",                 "Application Status",      "Current status of the onboarding application",              "No change needed",                                                                                                        "Application Status",       "Current status of the onboarding application"),
    ("change",  "application", "$native.application.subStatus",              "Application Sub Status",  "Detailed sub status of the application",                     "Adding hyphen improves readability. Description can be more informative.",                                             "Application Sub-Status",   "A more detailed status indicating the current stage within the application's status"),
    ("change",  "application", "$native.application.rejectionRemark",        "Rejection Remark",        "Reason or remark provided for application rejection",         "Minor: improving natural language flow.",                                                                                  "Rejection Remark",         "Reason or remark provided when the application was rejected"),
    ("none",    "application", "$native.application.dateCreated",            "Date Created",            "Timestamp when the application was created",                 "No change needed",                                                                                                        "Date Created",             "Timestamp when the application was created"),
    ("none",    "application", "$native.application.lastUpdated",            "Last Updated",            "Timestamp when the application was last updated",            "No change needed",                                                                                                        "Last Updated",             "Timestamp when the application was last updated"),
    ("none",    "application", "$native.application.loanPurpose",            "Loan Purpose",            "Purpose for which the loan is requested",                    "No change needed",                                                                                                        "Loan Purpose",             "Purpose for which the loan is requested"),
    ("none",    "application", "$native.application.appointmentDate",        "Appointment Date",        "Scheduled appointment date for the application",             "No change needed",                                                                                                        "Appointment Date",         "Scheduled appointment date for the application"),
    ("change",  "application", "$native.application.applicationSource",      "Application Source",      "Source from which the application was initiated",            "Minor: adding 'channel or' for clarity.",                                                                                 "Application Source",       "The channel or source from which the application was initiated"),
    ("none",    "application", "$native.application.lastStatusUpdated",      "Last Status Updated",     "Timestamp when the application status was last updated",     "No change needed",                                                                                                        "Last Status Updated",      "Timestamp when the application status was last updated"),

    # ── REDIRECT ATTEMPT ──────────────────────────────────────────────────────
    ("section", "REDIRECT ATTEMPT", "", "", "", "", "", ""),
    ("none",    "redirectAttempt", "$native.redirectAttempt.id",                    "Redirect Attempt ID",      "Unique identifier of the redirect attempt",                 "No change needed",                                                                                                  "Redirect Attempt ID",      "Unique identifier of the redirect attempt"),
    ("change",  "redirectAttempt", "$native.redirectAttempt.tenantId",              "Tenant ID",                "Tenant ID associated with the redirect attempt",             "Description restates the display name — adds no new information. Display name updated to match business terminology.", "Lender ID",                "Unique ID of the lender this redirect attempt is associated with"),
    ("change",  "redirectAttempt", "$native.redirectAttempt.applicantId",           "Applicant ID",             "Reference ID of the applicant",                              "Minor: adding context.",                                                                                            "Applicant ID",             "Reference ID of the applicant associated with this redirect attempt"),
    ("change",  "redirectAttempt", "$native.redirectAttempt.onboardingApplicationId","Onboarding Application ID","Reference ID of the onboarding application",               "Minor: adding context.",                                                                                            "Onboarding Application ID","Reference ID of the onboarding application for this redirect attempt"),
    ("none",    "redirectAttempt", "$native.redirectAttempt.redirectIdentifier",    "Redirect Identifier",      "External identifier used for the redirect attempt",          "No change needed",                                                                                                  "Redirect Identifier",      "External identifier used for the redirect attempt"),
    ("none",    "redirectAttempt", "$native.redirectAttempt.internalRedirectIdentifier","Internal Redirect Identifier","Internal system identifier for the redirect attempt", "No change needed",                                                                                                  "Internal Redirect Identifier","Internal system identifier for the redirect attempt"),
    ("change",  "redirectAttempt", "$native.redirectAttempt.type",                  "Redirect Attempt Type",    "Type of the redirect attempt",                              "Description restates the display name — adds no new information.",                                               "Redirect Attempt Type",    "The category or type of this redirect attempt"),
    ("none",    "redirectAttempt", "$native.redirectAttempt.status",                "Redirect Attempt Status",  "Current status of the redirect attempt",                     "No change needed",                                                                                                  "Redirect Attempt Status",  "Current status of the redirect attempt"),
    ("change",  "redirectAttempt", "$native.redirectAttempt.metadata",              "Metadata",                 "Additional metadata associated with the redirect attempt",   "Display name 'Metadata' is too generic — no table context.",                                                    "Redirect Attempt Metadata","Additional technical details associated with the redirect attempt"),
    ("change",  "redirectAttempt", "$native.redirectAttempt.expiryTime",            "Expiry Time",              "Time at which the redirect attempt expires",                 "Minor: 'Date and time' is more user-friendly than just 'Time'.",                                                 "Expiry Time",              "Date and time at which this redirect attempt expires"),
    ("none",    "redirectAttempt", "$native.redirectAttempt.dateCreated",            "Date Created",             "Timestamp when the redirect attempt was created",            "No change needed",                                                                                                  "Date Created",             "Timestamp when the redirect attempt was created"),
    ("none",    "redirectAttempt", "$native.redirectAttempt.lastUpdated",            "Last Updated",             "Timestamp when the redirect attempt was last updated",       "No change needed",                                                                                                  "Last Updated",             "Timestamp when the redirect attempt was last updated"),

    # ── USER ACTION ───────────────────────────────────────────────────────────
    ("section", "USER ACTION", "", "", "", "", "", ""),
    ("none",    "userAction", "$native.userAction.id",                    "User Action ID",           "Unique identifier of the user action",                       "No change needed",                                                                                                          "User Action ID",           "Unique identifier of the user action"),
    ("change",  "userAction", "$native.userAction.tenantId",              "Tenant ID",                "Tenant ID associated with the user action",                   "Description restates the display name — adds no new information. Display name updated to match business terminology.",    "Lender ID",                "Unique ID of the lender this user action is associated with"),
    ("change",  "userAction", "$native.userAction.applicantId",           "Applicant ID",             "Reference ID of the applicant",                               "Minor: adding context.",                                                                                                    "Applicant ID",             "Reference ID of the applicant who performed this action"),
    ("change",  "userAction", "$native.userAction.onboardingApplicationId","Onboarding Application ID","Reference ID of the onboarding application",                 "Minor: adding context.",                                                                                                    "Onboarding Application ID","Reference ID of the onboarding application for this user action"),
    ("change",  "userAction", "$native.userAction.applicantType",         "Applicant Type",           "Type of the applicant",                                       "Description too brief — restates the display name with no added clarity.",                                               "Applicant Type",           "The classification of the applicant who performed the action (e.g. primary applicant, co-applicant)"),
    ("change",  "userAction", "$native.userAction.actionType",            "Action Type",              "Type of action performed by the user",                        "Minor: 'user' should be 'applicant' for consistency.",                                                                    "Action Type",              "Type of action performed by the applicant"),
    ("change",  "userAction", "$native.userAction.actionTimeStamp",       "Action Timestamp",         "Timestamp when the user action occurred",                     "Minor: improving natural language.",                                                                                        "Action Timestamp",         "Date and time when the action was performed by the applicant"),
    ("none",    "userAction", "$native.userAction.applicantIpAddress",    "Applicant IP Address",     "IP address from which the applicant performed the action",    "No change needed",                                                                                                          "Applicant IP Address",     "IP address from which the applicant performed the action"),
    ("change",  "userAction", "$native.userAction.applicantHost",         "Applicant Host",           "Host details of the applicant's request source",              "'Applicant Host' and 'Host details' are both technical terms — not clear to a CST user.",                               "Applicant Device Host",    "Device or host information from which the applicant submitted the action"),
    ("change",  "userAction", "$native.userAction.metadata",              "Metadata",                 "Additional metadata associated with the user action",         "Display name 'Metadata' is too generic — no table context.",                                                              "User Action Metadata",     "Additional technical details associated with the user action"),
    ("none",    "userAction", "$native.userAction.dateCreated",           "Date Created",             "Timestamp when the user action record was created",           "No change needed",                                                                                                          "Date Created",             "Timestamp when the user action record was created"),
]

# ── Write headers ────────────────────────────────────────────────────────────
ws.append(HEADERS)
for col_idx, _ in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_wrap
    cell.border = thin_border

ws.row_dimensions[1].height = 30

# ── Write data rows ───────────────────────────────────────────────────────────
current_row = 2
for row_data in ROWS:
    row_type = row_data[0]
    values   = list(row_data[1:])  # table, key, curr_disp, curr_desc, comments, new_disp, new_desc

    if row_type == "section":
        ws.append([""] * 7)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        cell = ws.cell(row=current_row, column=1)
        cell.value = values[0]
        cell.fill  = section_fill
        cell.font  = section_font
        cell.alignment = Alignment(wrap_text=False, vertical="center", horizontal="left", indent=1)
        cell.border = thin_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1
        continue

    ws.append(values)

    # Determine row fill
    if row_type == "error":
        row_fill = error_fill
        row_font = error_font
    elif row_type == "pending":
        row_fill = pending_fill
        row_font = pending_font
    elif row_type == "change":
        row_fill = nochange_fill
        row_font = body_font
    else:  # none
        row_fill = nochange_fill
        row_font = body_font

    for col_idx in range(1, 8):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.border = thin_border
        cell.alignment = wrap

        if row_type in ("error", "pending"):
            cell.fill = row_fill
            cell.font = row_font
        else:
            cell.fill = nochange_fill
            cell.font = body_font
            # Green highlight only on new display name (col 6) and new description (col 7) when changed
            if row_type == "change" and col_idx in (6, 7):
                cell.fill = changed_fill

    ws.row_dimensions[current_row].height = 50
    current_row += 1

# ── Column widths ─────────────────────────────────────────────────────────────
col_widths = [18, 38, 26, 42, 52, 30, 52]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# ── Freeze panes ─────────────────────────────────────────────────────────────
ws.freeze_panes = "A2"

# ── Legend sheet ─────────────────────────────────────────────────────────────
ls = wb.create_sheet("Legend")
legend_data = [
    ("Color", "Meaning"),
    ("Red row",    "ERROR — display name or description is factually wrong or missing"),
    ("Yellow row", "PENDING CLARIFICATION — cannot finalize without your answer"),
    ("Green cells (col 6 & 7)", "CHANGED — new display name / description suggested"),
    ("White row",  "NO CHANGE — current values are clear and intuitive"),
]
for r_idx, (col_a, col_b) in enumerate(legend_data, 1):
    ls.cell(row=r_idx, column=1, value=col_a)
    ls.cell(row=r_idx, column=2, value=col_b)
    if r_idx == 1:
        ls.cell(row=r_idx, column=1).font = Font(bold=True)
        ls.cell(row=r_idx, column=2).font = Font(bold=True)
    fills = [None, error_fill, pending_fill, changed_fill, nochange_fill]
    if r_idx > 1:
        ls.cell(row=r_idx, column=1).fill = fills[r_idx - 1]

ls.column_dimensions["A"].width = 32
ls.column_dimensions["B"].width = 60

# ── Save ──────────────────────────────────────────────────────────────────────
output_path = r"e:\Antigravity\my-project\native_fields_review.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
